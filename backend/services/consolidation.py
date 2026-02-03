import pandas as pd
from typing import List, Dict
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def consolidate_timesheets(entries: List[Dict], project_name: str) -> bytes:
    """
    Consolidate timesheet entries by consultant and date.
    Returns Excel file as bytes.
    """
    # Convert to DataFrame
    # Convert to DataFrame
    df = pd.DataFrame(entries)
    
    # Fill missing project names with parent project name
    if 'project_name' not in df.columns:
        df['project_name'] = project_name
    else:
        df['project_name'] = df['project_name'].fillna(project_name)

    # Fill missing companies
    if 'company' not in df.columns:
        df['company'] = 'Unbekannt'
    else:
        df['company'] = df['company'].fillna('Unbekannt')
    
    # Group by keys including project_name
    # Group by: project, consultant, date, process, company
    group_keys = ['project_name', 'consultant_name', 'service_date', 'process_stream', 'company']
    
    # Aggregate
    consolidated = df.groupby(group_keys, dropna=False).agg({
        'hours': 'sum',
        'description': lambda x: ' | '.join(x.dropna().astype(str))
    }).reset_index()
    
    # Sort by Project, then Consultant, then Date
    consolidated = consolidated.sort_values(['project_name', 'consultant_name', 'service_date'])
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Konsolidierung"
    
    # --- Styles ---
    # Colors
    COLOR_PRIMARY = "1E293B"  # Slate 800 (Dark Blue/Grey)
    COLOR_ACCENT = "F8FAFC"   # Slate 50 (Very light grey for alternates/headers)
    COLOR_HEADER_BG = "E2E8F0" # Slate 200
    COLOR_TEXT = "0F172A"     # Slate 900
    COLOR_BORDER = "CBD5E1"   # Slate 300
    
    # Fonts
    font_title = Font(name="Calibri", size=14, bold=True, color=COLOR_PRIMARY)
    font_header = Font(name="Calibri", size=11, bold=True, color=COLOR_TEXT)
    font_cell = Font(name="Calibri", size=11, color=COLOR_TEXT)
    font_total = Font(name="Calibri", size=11, bold=True, color=COLOR_PRIMARY)
    
    # Fills
    fill_header = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Borders
    border_thin = Side(style='thin', color=COLOR_BORDER)
    border_double = Side(style='double', color=COLOR_BORDER)
    
    border_all = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    border_bottom_thick = Border(bottom=Side(style='medium', color=COLOR_BORDER))
    border_total = Border(top=border_double, bottom=border_thin)

    # Alignment
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')

    # --- Helper Function for Sheet Generation ---
    def create_grouped_sheet(wb, sheet_title, data_df, group_col, main_title):
        ws = wb.create_sheet(title=sheet_title) if sheet_title else wb.active
        ws.title = sheet_title if sheet_title else "Sheet1"
        
        # Global Title
        ws.merge_cells('A1:I1') # Extended range for new columns
        title_cell = ws['A1']
        title_cell.value = main_title
        title_cell.font = Font(name="Calibri", size=18, bold=True, color=COLOR_PRIMARY)
        title_cell.alignment = align_center
        
        row_idx = 3 # Start below title

        # Ensure group column exists
        if group_col not in data_df.columns:
            data_df[group_col] = "Unbekannt"

        # Get unique groups and iterate
        groups = data_df[group_col].unique()
        
        # Sort groups (handle None)
        try:
           groups = sorted(groups, key=lambda x: str(x) if x is not None else "")
        except:
           pass
        
        for group_name in groups:
            # Filter for current group
            group_df = data_df[data_df[group_col] == group_name]
            
            # 1. Group Header Section
            ws.merge_cells(f'A{row_idx}:I{row_idx}')
            group_header = ws.cell(row=row_idx, column=1)
            group_header.value = str(group_name)
            group_header.font = font_title
            group_header.border = border_bottom_thick
            row_idx += 1
            
            # 2. Column Headers
            for col_idx, col_cfg in enumerate(columns_config, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = col_cfg["header"]
                cell.font = font_header
                cell.fill = fill_header
                cell.border = border_all
                cell.alignment = align_center
            row_idx += 1
            
            # 3. Data Rows
            start_row = row_idx
            
            # Sort by date/consultant
            group_df = group_df.sort_values(['service_date', 'consultant_name'])
            
            for _, entry in group_df.iterrows():
                # Firma
                c1 = ws.cell(row=row_idx, column=1, value=entry.get('company', ''))
                c1.font = font_cell
                c1.border = border_all
                c1.alignment = align_left
                
                # Berater
                c2 = ws.cell(row=row_idx, column=2, value=entry.get('consultant_name', ''))
                c2.font = font_cell
                c2.border = border_all
                c2.alignment = align_left
                
                # Datum
                c3 = ws.cell(row=row_idx, column=3, value=entry.get('service_date', ''))
                c3.font = font_cell
                c3.border = border_all
                c3.alignment = align_center
                
                # Project
                c4 = ws.cell(row=row_idx, column=4, value=entry.get('project_name', ''))
                c4.font = font_cell
                c4.border = border_all
                c4.alignment = align_left

                # Phase
                c5 = ws.cell(row=row_idx, column=5, value=entry.get('project_phase', ''))
                c5.font = font_cell
                c5.border = border_all
                c5.alignment = align_left
                
                # Process
                c6 = ws.cell(row=row_idx, column=6, value=entry.get('process_stream', ''))
                c6.font = font_cell
                c6.border = border_all
                c6.alignment = align_left
                
                # Description
                # Description
                c7 = ws.cell(row=row_idx, column=7, value=entry.get('description', ''))
                c7.font = font_cell
                c7.border = border_all
                c7.alignment = align_left
                
                # Hours
                hours = entry.get('hours', 0.0)
                c8 = ws.cell(row=row_idx, column=8, value=hours)
                c8.font = font_cell
                c8.border = border_all
                c8.alignment = align_right
                c8.number_format = '0.00'
                
                # Non-billable
                nb_hours = entry.get('non_billable_hours', 0.0)
                c9 = ws.cell(row=row_idx, column=9, value=nb_hours if nb_hours else '')
                c9.font = font_cell
                c9.border = border_all
                c9.alignment = align_right
                c9.number_format = '0.00'
                
                row_idx += 1
                
            end_row = row_idx - 1 # Last data row
                
            # 4. Total Row
            # Label "Summe"
            ws.merge_cells(f'A{row_idx}:G{row_idx}')
            sum_label = ws.cell(row=row_idx, column=1)
            sum_label.value = f"Summe {group_name}"
            sum_label.font = font_total
            sum_label.alignment = Alignment(horizontal='right', vertical='center')
            sum_label.border = border_total
            
            # Value Formula Hours
            sum_val = ws.cell(row=row_idx, column=8)
            sum_val.value = f"=SUM(H{start_row}:H{end_row})"
            sum_val.font = font_total
            sum_val.alignment = align_right
            sum_val.border = border_total
            sum_val.number_format = '0.00'
            
            # Value Formula Non-Billable
            sum_val_nb = ws.cell(row=row_idx, column=9)
            sum_val_nb.value = f"=SUM(I{start_row}:I{end_row})"
            sum_val_nb.font = font_total
            sum_val_nb.alignment = align_right
            sum_val_nb.border = border_total
            sum_val_nb.number_format = '0.00'
            
            row_idx += 3 # Spacer
            
        # --- Final touches (Widths) ---
        for i, col_cfg in enumerate(columns_config, 1):
            col_letter = chr(64 + i) # A, B, C...
            ws.column_dimensions[col_letter].width = col_cfg["width"]

    # --- Setup Columns ---
    # We define columns dynamically based on sheet type in the loop above, but layout is similar
    # Columns map: 1:Company, 2:Consultant, 3:Date, 4:Project/Phase, 5:Process, 6:Desc, 7:Hours, 8:NonBillable
    columns_config = [
        {"header": "Firma", "key": "company", "width": 25},
        {"header": "Berater", "key": "consultant_name", "width": 20},
        {"header": "Datum", "key": "service_date", "width": 12},
        {"header": "Projekt", "key": "project_name", "width": 30},
        {"header": "Phase", "key": "project_phase", "width": 15},
        {"header": "IPC Prozess", "key": "process_stream", "width": 25},
        {"header": "Beschreibung", "key": "description", "width": 40},
        {"header": "Stunden", "key": "hours", "width": 12},
        {"header": "Nicht Verr.", "key": "non_billable_hours", "width": 12},
    ]

    # --- Execute ---
    # Ensure columns exist
    if 'project_phase' not in consolidated.columns: consolidated['project_phase'] = None
    if 'non_billable_hours' not in consolidated.columns: consolidated['non_billable_hours'] = 0.0
    if 'project_name' not in consolidated.columns: consolidated['project_name'] = project_name

    # Sheet 1: Nach Projekt
    create_grouped_sheet(wb, "Nach Projekt", consolidated, 'project_name', f"Leistungsnachweis: {project_name}")
    
    # Sheet 2: Nach Firma
    create_grouped_sheet(wb, "Nach Firma", consolidated, 'company', f"Leistungsnachweis nach Firma: {project_name}")

    # Remove default sheet if we created ours separately
    # The default sheet might be named "Sheet" or "Konsolidierung" if customized upstream
    for sheet_name in ["Sheet", "Konsolidierung"]:
        if sheet_name in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb[sheet_name]
    
    # Save to bytes
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()
